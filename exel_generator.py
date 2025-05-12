# This Python file uses the following encoding: utf-8
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_excel_report(json_file, output_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Инфо объекты"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    group_font = Font(bold=True, size=12, color="FFFFFF")
    group_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers
    headers = [
        "IO Index",
        "Наименование (рус)",
        "Наименование (англ)",
        "Тип данных",
        "Используется в логике",
        "Сохранение",
        "Уставка апертуры",
        "КТТ",
        "Значение по умолчанию"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment

    # Group data by paramType
    groups = {
        "Аналоговые входы": "Входные сигналы",
        "Дискретные входы": "Входные сигналы",
        "Аналоговый выход": "Выходные сигналы",
        "Дискретный выход": "Выходные сигналы",
        "Признаки": "Признаки",
        "Уставка": "Уставки"
    }

    grouped_data = {}
    for item in data:
        group_name = groups.get(item["paramType"], "Другие")
        if group_name not in grouped_data:
            grouped_data[group_name] = []
        grouped_data[group_name].append(item)

    # Write data to Excel
    current_row = 2

    for group_name, items in grouped_data.items():
        # Write group header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        group_cell = ws.cell(row=current_row, column=1, value=group_name)
        group_cell.font = group_font
        group_cell.fill = group_fill
        group_cell.border = thin_border
        group_cell.alignment = center_alignment
        current_row += 1

        # Write items in group
        for item in items:
            ws.cell(row=current_row, column=1, value=item.get("ioIndex", "")).border = thin_border
            ws.cell(row=current_row, column=2, value=item.get("name", "").strip()).border = thin_border
            ws.cell(row=current_row, column=3, value=item.get("codeName", "").strip()).border = thin_border
            ws.cell(row=current_row, column=4, value=item.get("type", "")).border = thin_border

            logicuse = item.get("logicuse", "")
            ws.cell(row=current_row, column=5, value="Да" if logicuse == "Да" else "").border = thin_border

            saving = item.get("saving", "")
            ws.cell(row=current_row, column=6, value="Да" if saving == "Да" else "").border = thin_border

            ws.cell(row=current_row, column=7, value=item.get("aperture", "")).border = thin_border
            ws.cell(row=current_row, column=8, value=item.get("ktt", "")).border = thin_border
            ws.cell(row=current_row, column=9, value=item.get("def_value", "")).border = thin_border

            current_row += 1

    column_widths = [10, 70, 40, 15, 30, 15, 15, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=current_row+1, column=1, value=f"{timestamp}")

    wb.save(output_file)
    print(f"Excel file saved as {output_file}")

if __name__ == "__main__":
    input_json = "export.json"  # Change to your input file
    output_excel = "Инфо объекты.xlsx"  # Change to desired output file
    create_excel_report(input_json, output_excel)
